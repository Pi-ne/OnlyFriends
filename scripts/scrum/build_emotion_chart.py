# -*- coding: utf-8 -*-
"""生成 OnlyFriends 团队情绪图 Excel。"""
from __future__ import annotations

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.data_source import NumData, NumRef, NumVal, StrData, StrRef, StrVal
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUT_PATH = os.path.join(ROOT, "docs", "product", "情绪图.xlsx")

members = [
    ("A", "产品负责人（PO）", "迭代目标对齐、优先级决策、评审协调"),
    ("B", "后端开发工程师", "User / Activity 等核心服务开发与接口实现"),
    ("C", "产品经理", "用户故事、验收标准、需求评审与故事地图"),
    ("D", "前端开发工程师", "微信小程序与管理后台页面开发"),
    ("E", "AI 工程师", "内容审核、智能体接入与 AI 服务兜底"),
    ("F", "基础设施工程师", "Gateway、部署环境、Docker 与中间件"),
    ("G", "后端代码审查工程师", "接口规范审查、代码评审与质量把关"),
    ("H", "技术文档工程师", "发布计划、用户手册、实践总结等文档"),
    ("I", "质量保障工程师", "测试计划、用例设计与缺陷跟踪"),
]

# 数据截止 2026-06-29（第 3 天）
days = [
    ("第0天", "2026-06-26", "项目启动"),
    ("第1天", "2026-06-27", "v0.0 需求基线"),
    ("第2天", "2026-06-28", "v0.0 文档与环境基线"),
    ("第3天", "2026-06-29", "v0.1 首次迭代启动"),
]

scores = {
    "A": [4, 3, 4, 4],
    "B": [4, 3, 2, 3],
    "C": [4, 3, 3, 4],
    "D": [4, 3, 3, 4],
    "E": [4, 4, 3, 4],
    "F": [4, 3, 2, 3],
    "G": [4, 3, 3, 4],
    "H": [4, 2, 3, 4],
    "I": [4, 3, 3, 4],
}

day_notes = [
    "启动会，团队组建，期待感强",
    "需求拆分与验收标准撰写，文档与故事工作量大",
    "开发环境搭建与工具培训，部分服务启动遇阻",
    "首次迭代计划会，任务拆分明确，情绪回升",
]


def score_fill(value: int) -> PatternFill:
    if value >= 4:
        return PatternFill("solid", fgColor="C6EFCE")
    if value == 3:
        return PatternFill("solid", fgColor="FFEB9C")
    return PatternFill("solid", fgColor="FFC7CE")


def _parse_range(formula: str, default_ws):
    sheet_name = default_ws.title
    range_part = formula
    if "!" in formula:
        sheet_part, range_part = formula.rsplit("!", 1)
        sheet_name = sheet_part.strip("'")
    ws = default_ws.parent[sheet_name]
    range_part = range_part.replace("$", "")
    start, end = range_part.split(":") if ":" in range_part else (range_part, range_part)
    sc, sr = coordinate_from_string(start)
    ec, er = coordinate_from_string(end)
    return ws, column_index_from_string(sc), sr, column_index_from_string(ec), er


def _fill_num_ref(num_ref: NumRef, default_ws) -> None:
    if num_ref is None or not num_ref.f:
        return
    ws, c1, r1, c2, r2 = _parse_range(num_ref.f, default_ws)
    pts: list[NumVal] = []
    idx = 0
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            if isinstance(value, datetime):
                num = to_excel(value)
            else:
                num = float(value)
            pts.append(NumVal(idx=idx, v=num))
            idx += 1
    num_ref.numCache = NumData(pt=pts, formatCode="General")


def _fill_str_ref(str_ref: StrRef, default_ws) -> None:
    if str_ref is None or not str_ref.f:
        return
    ws, c1, r1, c2, r2 = _parse_range(str_ref.f, default_ws)
    pts: list[StrVal] = []
    idx = 0
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            pts.append(StrVal(idx=idx, v=str(value)))
            idx += 1
    str_ref.strCache = StrData(pt=pts)


def _populate_chart_caches(ws, chart: LineChart) -> None:
    for series in chart.series:
        if series.val and series.val.numRef:
            _fill_num_ref(series.val.numRef, ws)
        if series.title and series.title.strRef:
            _fill_str_ref(series.title.strRef, ws)
        cat = series.cat
        if cat is None:
            continue
        if cat.strRef:
            _fill_str_ref(cat.strRef, ws)
        elif cat.numRef:
            _fill_num_ref(cat.numRef, ws)


def _add_trend_chart_sheet(wb: Workbook, ws_data, avg_row: int) -> None:
    """在独立工作表生成可正确渲染的折线图（避免按列误解析导致空白）。"""
    ws_chart = wb.create_sheet("情绪趋势图", 1)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center", vertical="center")

    ws_chart["A1"] = "情绪趋势图数据源（随「情绪数据」表同步更新）"
    ws_chart["A1"].font = Font(bold=True, size=12)

    ws_chart.cell(2, 1, "成员").font = header_font
    ws_chart.cell(2, 1).fill = header_fill
    ws_chart.cell(2, 1).alignment = center
    for j, (day_label, date_str, _) in enumerate(days, start=2):
        cell = ws_chart.cell(2, j, datetime.strptime(date_str, "%Y-%m-%d"))
        cell.number_format = "mm-dd"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    chart_rows: list[tuple[str, list]] = [("团队平均", [])]
    for j in range(len(days)):
        col = j + 3
        chart_rows[0][1].append(ws_data.cell(avg_row, col).value)
    for mid, _, _ in members:
        row_vals = [scores[mid][j] for j in range(len(days))]
        chart_rows.append((mid, row_vals))

    for i, (name, vals) in enumerate(chart_rows, start=3):
        ws_chart.cell(i, 1, name).alignment = center
        if name == "团队平均":
            ws_chart.cell(i, 1).font = Font(bold=True)
        for j, val in enumerate(vals, start=2):
            c = ws_chart.cell(i, j, val)
            c.alignment = center
            if name == "团队平均":
                c.font = Font(bold=True)

    ws_chart.column_dimensions["A"].width = 10
    for j in range(len(days)):
        ws_chart.column_dimensions[get_column_letter(2 + j)].width = 12

    chart = LineChart()
    chart.title = "OnlyFriends 团队情绪趋势图"
    chart.style = 10
    chart.y_axis.title = "情绪分值"
    chart.x_axis.title = "日期"
    chart.y_axis.scaling.min = 1
    chart.y_axis.scaling.max = 5
    chart.height = 12
    chart.width = 20

    first_data_row = 3
    last_data_row = first_data_row + len(chart_rows) - 1
    first_val_col = 2
    last_val_col = first_val_col + len(days) - 1

    cats = Reference(
        ws_chart,
        min_col=first_val_col,
        min_row=2,
        max_col=last_val_col,
        max_row=2,
    )
    data = Reference(
        ws_chart,
        min_col=1,
        min_row=first_data_row,
        max_col=last_val_col,
        max_row=last_data_row,
    )
    chart.add_data(data, from_rows=True, titles_from_data=True)
    chart.set_categories(cats)
    _populate_chart_caches(ws_chart, chart)
    chart.legend.position = "b"
    ws_chart.add_chart(chart, "A16")


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "情绪数据"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    last_col = 2 + len(days)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = "OnlyFriends 团队情绪图（成员 A-I）"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws["A2"] = "评分说明：5=非常开心  4=不错/期待  3=平静  2=有压力  1=沮丧    |    数据截止：2026-06-29（第 3 天）"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    row_hdr = 4
    ws.cell(row_hdr, 1, "成员").font = header_font
    ws.cell(row_hdr, 1).fill = header_fill
    ws.cell(row_hdr, 1).alignment = center
    ws.cell(row_hdr, 1).border = border

    ws.cell(row_hdr, 2, "角色").font = header_font
    ws.cell(row_hdr, 2).fill = header_fill
    ws.cell(row_hdr, 2).alignment = center
    ws.cell(row_hdr, 2).border = border

    for j, (day_label, date_str, _milestone) in enumerate(days, start=3):
        cell = ws.cell(row_hdr, j, f"{day_label}\n{date_str}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    note_col = 3 + len(days)
    ws.cell(row_hdr, note_col, "角色分工").font = header_font
    ws.cell(row_hdr, note_col).fill = header_fill
    ws.cell(row_hdr, note_col).alignment = center
    ws.cell(row_hdr, note_col).border = border

    for i, (mid, role, duty) in enumerate(members, start=row_hdr + 1):
        ws.cell(i, 1, mid).alignment = center
        ws.cell(i, 1).font = Font(bold=True)
        ws.cell(i, 1).border = border
        ws.cell(i, 2, role).alignment = center
        ws.cell(i, 2).border = border
        for j, sc in enumerate(scores[mid], start=3):
            c = ws.cell(i, j, sc)
            c.alignment = center
            c.border = border
            c.fill = score_fill(sc)
        ws.cell(i, note_col, duty).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(i, note_col).border = border

    avg_row = row_hdr + 1 + len(members)
    ws.cell(avg_row, 1, "团队平均").font = Font(bold=True)
    ws.cell(avg_row, 1).alignment = center
    ws.cell(avg_row, 1).border = border
    ws.cell(avg_row, 2, "-").alignment = center
    ws.cell(avg_row, 2).border = border
    for j in range(len(days)):
        col = j + 3
        vals = [scores[m[0]][j] for m in members]
        avg = round(sum(vals) / len(vals), 2)
        c = ws.cell(avg_row, col, avg)
        c.font = Font(bold=True)
        c.alignment = center
        c.border = border
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(avg_row, note_col, "每日加权平均（满分5）").alignment = center
    ws.cell(avg_row, note_col).border = border

    note_row = avg_row + 1
    ws.cell(note_row, 1, "当日备注").font = Font(bold=True, italic=True)
    ws.cell(note_row, 1).alignment = center
    for j, note in enumerate(day_notes, start=3):
        c = ws.cell(note_row, j, note)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=9, italic=True, color="666666")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    for j in range(len(days)):
        ws.column_dimensions[get_column_letter(3 + j)].width = 11
    ws.column_dimensions[get_column_letter(note_col)].width = 28

    _add_trend_chart_sheet(wb, ws, avg_row)

    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ["OnlyFriends 情绪图使用说明", ""],
        ["", ""],
        ["1. 评分标准", ""],
        ["分值", "含义"],
        ["5", "非常开心 / 元气满满 — 进展顺利、信心足"],
        ["4", "不错 / 充满期待 — 状态良好"],
        ["3", "平静 / 一般 — 正常节奏，无明显波动"],
        ["2", "有压力 / 有点累 — 遇到阻塞或任务偏重"],
        ["1", "沮丧 / 压力山大 — 需要团队介入支持"],
        ["", ""],
        ["2. 更新频率", "每日站会结束后收集各成员自评并更新本表"],
        ["3. 数据范围", "本表截至 2026-06-29（第 3 天），后续日程随迭代推进逐日追加"],
        ["4. 预警规则", "团队平均分连续 2 天低于 2.5 时，在回顾会中讨论阻塞项与分工调整"],
        [
            "5. 成员代号",
            "A=PO  B=后端开发  C=产品经理  D=前端开发  E=AI工程师  "
            "F=基础设施  G=后端代码审查  H=技术文档  I=质量保障",
        ],
        ["", ""],
        ["6. 已记录日程与情绪关联", ""],
        ["日程", "版本", "典型情绪变化"],
        ["第0天 06-26", "启动", "期待感高，整体 4 分左右"],
        ["第1天 06-27", "v0.0", "故事与文档撰写密集，产品/文档角色压力偏大"],
        ["第2天 06-28", "v0.0", "环境搭建遇阻，后端与基础设施角色偏低"],
        ["第3天 06-29", "v0.1 启动", "迭代计划明确，整体情绪回升"],
    ]
    for r, row in enumerate(instructions, 1):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(r, c, val)
            if r in (1, 3, 16):
                cell.font = Font(bold=True, size=12 if r == 1 else 11)
            if r == 4:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E7E6E6")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 55

    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Size: {os.path.getsize(OUT_PATH)} bytes")


if __name__ == "__main__":
    main()
