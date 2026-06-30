# -*- coding: utf-8 -*-
"""生成 OnlyFriends 项目看板 Excel。"""
from __future__ import annotations

import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUT_PATH = os.path.join(ROOT, "docs", "product", "看板.xlsx")

# 数据截止 2026-06-29（第 3 天），与情绪图 / 燃尽图日程一致
CUTOFF = "2026-06-29"
SPRINT = "v0.1 首次迭代（06-29 启动）"

# (编号, 标题, 负责人, 理想时, 优先级, 版本, 状态, 备注)
# 状态：待办 | 进行中 | 阻塞 | 已完成
WORK_ITEMS = [
    # v0.0 已完成
    ("T01", "团队角色表与运行规范", "A", 6, "P0", "v0.0", "已完成", "第0天完成"),
    ("T02", "原始需求分析概要", "C", 8, "P0", "v0.0", "已完成", "第1天完成"),
    ("T03", "用户故事清单初稿", "C", 12, "P0", "v0.0", "已完成", "第1天完成"),
    ("T04", "验收标准文档", "C,I", 16, "P0", "v0.0", "已完成", "第1天完成"),
    ("T05", "版本发布计划初稿", "A,H", 10, "P0", "v0.0", "已完成", "第1天完成"),
    ("T06", "故事地图与优先级排序", "C", 8, "P0", "v0.0", "已完成", "第2天完成"),
    ("T07", "用户故事与 AC 评审优化", "C", 10, "P0", "v0.0", "已完成", "第2天完成"),
    ("T08", "后端工程本地编译验证", "B,G", 8, "P0", "v0.0", "已完成", "第2天完成"),
    ("T11", "Gateway 路由与鉴权骨架验证", "F,I", 6, "P0", "v0.0", "已完成", "第2天完成"),
    ("T12", "首次迭代计划文档", "A", 8, "P0", "v0.1", "已完成", "第3天完成"),
    ("T13", "故事到任务拆分（v0.1 P0）", "A,C", 10, "P0", "v0.1", "已完成", "第3天完成"),
    ("T14", "理想时估算与看板/燃尽图建立", "全组", 8, "P0", "v0.1", "已完成", "第3天完成"),
    # v0.0 进行中 / 阻塞
    ("T09", "Docker Compose 与中间件启动", "F", 14, "P0", "v0.0", "进行中", "Redis/MinIO 已启动，联调待验证"),
    ("T10", "数据库初始化与连接配置", "B,F", 8, "P0", "v0.0", "进行中", "脚本可执行，IDE 直连待统一密码"),
    ("T10-B", "本地 DB 密码与 application.yml 对齐", "F", 2, "P0", "v0.0", "阻塞", "需 IDE 运行配置引用 set-local-env.ps1"),
    # v0.1 Sprint Backlog（待办）
    ("K01", "US-U-001 用户注册（接口 + 小程序页）", "B,D", 8, "P0", "v0.1", "待办", "主链路起点"),
    ("K02", "US-U-002 账号激活（模拟激活）", "B", 4, "P0", "v0.1", "待办", ""),
    ("K03", "US-U-003 登录与 Token 签发", "B,D", 8, "P0", "v0.1", "待办", ""),
    ("K04", "US-U-004 Access Token 自动刷新", "B,D", 4, "P0", "v0.1", "待办", ""),
    ("K05", "US-A-001 活动创建与草稿保存", "B,D", 10, "P0", "v0.1", "待办", ""),
    ("K06", "US-A-002 活动提交 AI/人工审核", "B,E", 8, "P0", "v0.1", "待办", "AI 可用 Mock 兜底"),
    ("K07", "US-A-003 活动列表浏览（首页）", "B,D", 8, "P0", "v0.1", "待办", ""),
    ("K08", "US-A-006 活动详情页", "D", 6, "P0", "v0.1", "待办", ""),
    ("K09", "US-A-009 活动报名", "B,D", 8, "P0", "v0.1", "待办", ""),
    ("K10", "US-A-010 报名状态查询", "B", 4, "P0", "v0.1", "待办", ""),
    ("K11", "US-AI-001 活动内容安全审核兜底", "E", 8, "P0", "v0.1", "待办", "规则引擎 / Mock"),
    ("K12", "US-ADM-001 管理员登录", "B,D", 4, "P0", "v0.1", "待办", "管理后台"),
    ("K13", "US-ADM-002 活动审核通过/驳回", "B,D", 8, "P0", "v0.1", "待办", "管理后台"),
    ("K14", "US-INF-001 Gateway 路由与鉴权联调", "F,B", 10, "P0", "v0.1", "待办", ""),
    ("K15", "US-INF-002 服务契约与统一响应校验", "B,G", 6, "P0", "v0.1", "待办", ""),
    ("K16", "主链路端到端联调", "全组", 12, "P0", "v0.1", "待办", "目标：06-30 打通"),
    ("K17", "P0 测试用例设计与冒烟脚本", "I", 8, "P1", "v0.1", "待办", ""),
    ("K18", "演示数据、部署说明与账号清单", "F,H", 6, "P1", "v0.1", "待办", "第一次评审材料"),
]

COLUMNS = ["待办", "进行中", "阻塞", "已完成"]
COLUMN_FILLS = {
    "待办": "FFF2CC",
    "进行中": "DDEBF7",
    "阻塞": "FCE4D6",
    "已完成": "E2EFDA",
}
COLUMN_HEADER_FILLS = {
    "待办": "BF8F00",
    "进行中": "2F75B5",
    "阻塞": "C55A11",
    "已完成": "548235",
}


def _card_text(item: tuple) -> str:
    wid, title, owner, hours, priority, version, _status, note = item
    lines = [f"[{wid}] {title}", f"{owner} | {hours}h | {priority} | {version}"]
    if note:
        lines.append(note)
    return "\n".join(lines)


def _thin_border() -> Border:
    thin = Side(style="thin", color="B4B4B4")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _add_kanban_view(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "看板视图"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    card_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = "OnlyFriends 项目看板"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"当前迭代：{SPRINT}  |  数据截止：{CUTOFF}（第 3 天）  |  "
        f"列：待办 → 进行中 → 阻塞 → 已完成"
    )
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    by_status: dict[str, list] = defaultdict(list)
    for item in WORK_ITEMS:
        by_status[item[6]].append(item)

    stats_row = 4
    ws.cell(stats_row, 1, "工作项统计").font = Font(bold=True)
    col = 2
    for status in COLUMNS:
        count = len(by_status[status])
        hours = sum(i[3] for i in by_status[status])
        cell = ws.cell(stats_row, col, f"{status}：{count} 项 / {hours}h")
        cell.fill = PatternFill("solid", fgColor=COLUMN_FILLS[status])
        cell.alignment = center
        cell.border = _thin_border()
        col += 1

    hdr_row = 6
    col_widths = [32, 32, 28, 32]
    for i, status in enumerate(COLUMNS):
        col_idx = i + 1
        cell = ws.cell(hdr_row, col_idx, status)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=COLUMN_HEADER_FILLS[status])
        cell.alignment = center
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[i]

    max_cards = max(len(by_status[s]) for s in COLUMNS)
    for row_offset in range(max_cards):
        row = hdr_row + 1 + row_offset
        for i, status in enumerate(COLUMNS):
            col_idx = i + 1
            items = by_status[status]
            cell = ws.cell(row, col_idx)
            cell.border = _thin_border()
            cell.alignment = top_wrap
            cell.font = card_font
            cell.fill = PatternFill("solid", fgColor=COLUMN_FILLS[status])
            if row_offset < len(items):
                cell.value = _card_text(items[row_offset])
                ws.row_dimensions[row].height = 58

    legend_row = hdr_row + max_cards + 2
    ws.cell(legend_row, 1, "图例").font = Font(bold=True)
    ws.cell(legend_row + 1, 1, "卡片格式：[编号] 任务标题 / 负责人 / 理想时 / 优先级")
    ws.cell(legend_row + 2, 1, "v0.0 规划任务已基本完成；v0.1 主链路开发任务已进入 Sprint Backlog（待办）")


def _add_work_item_list(wb: Workbook) -> None:
    ws = wb.create_sheet("工作项列表")

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = _thin_border()

    headers = ["编号", "标题", "负责人", "理想时(h)", "优先级", "版本", "状态", "备注"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(1, col, text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    status_fill = {
        "待办": PatternFill("solid", fgColor="FFF2CC"),
        "进行中": PatternFill("solid", fgColor="DDEBF7"),
        "阻塞": PatternFill("solid", fgColor="FCE4D6"),
        "已完成": PatternFill("solid", fgColor="E2EFDA"),
    }

    for row_idx, item in enumerate(WORK_ITEMS, start=2):
        for col_idx, val in enumerate(item, start=1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.alignment = wrap if col_idx == 2 else center
            cell.border = border
            if col_idx == 7:
                cell.fill = status_fill.get(val, PatternFill())

    widths = [10, 36, 10, 10, 8, 8, 10, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def _add_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("填写说明")

    instructions = [
        ["OnlyFriends 看板使用说明", ""],
        ["", ""],
        ["1. 看板列定义", ""],
        ["列名", "含义"],
        ["待办", "已纳入当前迭代但尚未开始的工作项（Sprint Backlog）"],
        ["进行中", "成员已认领并正在执行的工作项"],
        ["阻塞", "因依赖、环境或决策未明而暂停的工作项"],
        ["已完成", "满足验收标准并经 PO 确认关闭的工作项"],
        ["", ""],
        ["2. 数据范围", f"本看板截至 {CUTOFF}（第 3 天），覆盖 v0.0 收尾与 v0.1 迭代计划结果"],
        ["3. 更新频率", "每日站会结束后由 PO 或 Scrum Master 更新卡片位置与备注"],
        ["4. WIP 限制", "建议进行中列同时不超过 4 项，避免并行过多导致联调阻塞"],
        ["5. 成员代号", "A=PO  B=后端  C=产品  D=前端  E=AI  F=基础设施  G=代码审查  H=文档  I=质量"],
        ["", ""],
        ["6. 截至 6.29 看板状态摘要", ""],
        ["类别", "说明"],
        ["v0.0 规划", "12 项已完成；环境搭建 2 项进行中；DB 配置 1 项阻塞"],
        ["v0.1 开发", "18 项主链路任务已进入待办，按 User→Activity→Admin→Gateway→联调顺序推进"],
        ["下一步", "06-30 优先解除 T10-B 阻塞，启动 K01~K04 用户认证主链路开发"],
        ["", ""],
        ["7. 关联文档", "情绪图.xlsx / 燃尽图.xlsx / release-plan.md / user-stories.md"],
    ]

    for r, row in enumerate(instructions, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if r in (1, 3, 6):
                cell.font = Font(bold=True, size=12 if r == 1 else 11)
            if r == 4:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E7E6E6")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 62


def build_workbook() -> Workbook:
    wb = Workbook()
    _add_kanban_view(wb)
    _add_work_item_list(wb)
    _add_instructions(wb)
    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Size: {os.path.getsize(OUT_PATH)} bytes")
    by_status = defaultdict(int)
    for item in WORK_ITEMS:
        by_status[item[6]] += 1
    print("Items:", dict(by_status))


if __name__ == "__main__":
    main()
