# -*- coding: utf-8 -*-
"""生成 OnlyFriends 项目燃尽图 Excel。"""
from __future__ import annotations

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.data_source import NumData, NumRef, NumVal, StrData, StrRef, StrVal
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils.datetime import to_excel

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUT_PATH = os.path.join(ROOT, "docs", "product", "燃尽图.xlsx")

# 数据截止 2026-06-29（第 3 天），与情绪图日程一致
days = [
    ("第0天", "2026-06-26", "项目启动"),
    ("第1天", "2026-06-27", "v0.0 需求基线"),
    ("第2天", "2026-06-28", "v0.0 文档与环境基线"),
    ("第3天", "2026-06-29", "v0.1 首次迭代启动"),
]

# v0.0 启动与规划阶段任务（理想时合计 132h）
tasks = [
    ("T01", "团队角色表与运行规范", 6, "A", 0, "已完成", "v0.0"),
    ("T02", "原始需求分析概要", 8, "C", 1, "已完成", "v0.0"),
    ("T03", "用户故事清单初稿", 12, "C", 1, "已完成", "v0.0"),
    ("T04", "验收标准文档", 16, "C,I", 1, "已完成", "v0.0"),
    ("T05", "版本发布计划初稿", 10, "A,H", 1, "已完成", "v0.0"),
    ("T06", "故事地图与优先级排序", 8, "C", 2, "已完成", "v0.0"),
    ("T07", "用户故事与 AC 评审优化", 10, "C", 2, "已完成", "v0.0"),
    ("T08", "后端工程本地编译验证", 8, "B,G", 2, "已完成", "v0.0"),
    ("T09", "Docker Compose 与中间件启动", 14, "F", 2, "进行中", "v0.0"),
    ("T10", "数据库初始化与连接配置", 8, "B,F", 2, "进行中", "v0.0"),
    ("T11", "Gateway 路由与鉴权骨架验证", 6, "F,I", 2, "已完成", "v0.0"),
    ("T12", "首次迭代计划文档", 8, "A", 3, "已完成", "v0.1"),
    ("T13", "故事到任务拆分（v0.1 P0）", 10, "A,C", 3, "已完成", "v0.1"),
    ("T14", "理想时估算与看板/燃尽图建立", 8, "全组", 3, "已完成", "v0.1"),
]

TOTAL_IDEAL_HOURS = sum(t[2] for t in tasks)

# 每日末剩余工时：期初 + 4 个日程节点
ideal_remaining = [TOTAL_IDEAL_HOURS, 99, 66, 33, 0]
actual_remaining = [TOTAL_IDEAL_HOURS, 126, 80, 42, 10]

day_completed = [
    TOTAL_IDEAL_HOURS - actual_remaining[1],
    actual_remaining[1] - actual_remaining[2],
    actual_remaining[2] - actual_remaining[3],
    actual_remaining[3] - actual_remaining[4],
]

day_notes = [
    "启动会，完成团队规范与角色对齐",
    "用户故事、验收标准、发布计划集中产出",
    "环境搭建遇阻，Docker/DB 配置滞后于计划",
    "迭代计划会完成，任务拆分明确，剩余 10h 转入 v0.1",
]


def _parse_range(formula: str, default_ws):
    sheet_name = default_ws.title
    range_part = formula
    if "!" in formula:
        sheet_part, range_part = formula.rsplit("!", 1)
        sheet_name = sheet_part.strip("'")
    ws = default_ws.parent[sheet_name]
    range_part = range_part.replace("$", "")
    start, end = range_part.split(":") if ":" in range_part else (range_part, range_part)
    sc, sr = coordinate_from_string(start)
    ec, er = coordinate_from_string(end)
    return ws, column_index_from_string(sc), sr, column_index_from_string(ec), er


def _fill_num_ref(num_ref: NumRef, default_ws) -> None:
    if num_ref is None or not num_ref.f:
        return
    ws, c1, r1, c2, r2 = _parse_range(num_ref.f, default_ws)
    pts: list[NumVal] = []
    idx = 0
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            if isinstance(value, datetime):
                num = to_excel(value)
            else:
                num = float(value)
            pts.append(NumVal(idx=idx, v=num))
            idx += 1
    num_ref.numCache = NumData(pt=pts, formatCode="General")


def _fill_str_ref(str_ref: StrRef, default_ws) -> None:
    if str_ref is None or not str_ref.f:
        return
    ws, c1, r1, c2, r2 = _parse_range(str_ref.f, default_ws)
    pts: list[StrVal] = []
    idx = 0
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            pts.append(StrVal(idx=idx, v=str(value)))
            idx += 1
    str_ref.strCache = StrData(pt=pts)


def _populate_chart_caches(ws, chart: LineChart) -> None:
    for series in chart.series:
        if series.val and series.val.numRef:
            _fill_num_ref(series.val.numRef, ws)
        if series.title and series.title.strRef:
            _fill_str_ref(series.title.strRef, ws)
        cat = series.cat
        if cat is None:
            continue
        if cat.strRef:
            _fill_str_ref(cat.strRef, ws)
        elif cat.numRef:
            _fill_num_ref(cat.numRef, ws)


def _status_fill(status: str) -> PatternFill:
    if status == "已完成":
        return PatternFill("solid", fgColor="C6EFCE")
    if status == "进行中":
        return PatternFill("solid", fgColor="FFEB9C")
    return PatternFill("solid", fgColor="FFC7CE")


def _add_burndown_chart_sheet(wb: Workbook) -> None:
    ws_chart = wb.create_sheet("燃尽趋势图", 1)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center", vertical="center")

    ws_chart["A1"] = "燃尽趋势图数据源（随「燃尽数据」表同步更新）"
    ws_chart["A1"].font = Font(bold=True, size=12)

    ws_chart.cell(2, 1, "曲线").font = header_font
    ws_chart.cell(2, 1).fill = header_fill
    ws_chart.cell(2, 1).alignment = center

    for j, (_, date_str, _) in enumerate(days, start=2):
        cell = ws_chart.cell(2, j, datetime.strptime(date_str, "%Y-%m-%d"))
        cell.number_format = "mm-dd"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 图表仅展示各日程结束时的剩余工时（不含期初点）
    chart_ideal = ideal_remaining[1:]
    chart_actual = actual_remaining[1:]
    chart_rows = [
        ("理想剩余工时", chart_ideal),
        ("实际剩余工时", chart_actual),
    ]
    for i, (name, vals) in enumerate(chart_rows, start=3):
        ws_chart.cell(i, 1, name).alignment = center
        ws_chart.cell(i, 1).font = Font(bold=True)
        for j, val in enumerate(vals, start=2):
            c = ws_chart.cell(i, j, val)
            c.alignment = center

    ws_chart.column_dimensions["A"].width = 16
    for j in range(len(days)):
        ws_chart.column_dimensions[get_column_letter(2 + j)].width = 12

    chart = LineChart()
    chart.title = "OnlyFriends 项目燃尽图（v0.0 启动与规划）"
    chart.style = 10
    chart.y_axis.title = "剩余理想时（小时）"
    chart.x_axis.title = "日期"
    chart.y_axis.scaling.min = 0
    chart.height = 12
    chart.width = 20

    first_data_row = 3
    last_data_row = first_data_row + len(chart_rows) - 1
    first_val_col = 2
    last_val_col = first_val_col + len(days) - 1

    cats = Reference(
        ws_chart,
        min_col=first_val_col,
        min_row=2,
        max_col=last_val_col,
        max_row=2,
    )
    data = Reference(
        ws_chart,
        min_col=1,
        min_row=first_data_row,
        max_col=last_val_col,
        max_row=last_data_row,
    )
    chart.add_data(data, from_rows=True, titles_from_data=True)
    chart.set_categories(cats)
    _populate_chart_caches(ws_chart, chart)
    chart.legend.position = "b"
    ws_chart.add_chart(chart, "A8")


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "燃尽数据"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    last_col = 2 + len(days)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = "OnlyFriends 项目燃尽图（v0.0 启动与规划阶段）"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws["A2"] = (
        f"范围：v0.0 交付物 + 第 3 天 v0.1 迭代计划基线  |  "
        f"总理想时：{TOTAL_IDEAL_HOURS}h  |  数据截止：2026-06-29（第 3 天）"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    row_hdr = 4
    headers = ["指标", "期初"] + [f"{d[0]}\n{d[1]}" for d in days]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row_hdr, col, text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    summary_rows = [
        ("理想剩余工时（h）", ideal_remaining),
        ("实际剩余工时（h）", actual_remaining),
        ("当日完成工时（h）", [None] + day_completed),
        (
            "累计完成率",
            [None]
            + [
                f"{round((TOTAL_IDEAL_HOURS - actual_remaining[i + 1]) / TOTAL_IDEAL_HOURS * 100, 1)}%"
                for i in range(len(days))
            ],
        ),
    ]

    for i, (label, vals) in enumerate(summary_rows, start=row_hdr + 1):
        ws.cell(i, 1, label).font = Font(bold=True)
        ws.cell(i, 1).alignment = center
        ws.cell(i, 1).border = border
        for j, val in enumerate(vals, start=2):
            c = ws.cell(i, j, val)
            c.alignment = center
            c.border = border
            if label.startswith("理想"):
                c.fill = PatternFill("solid", fgColor="D9E1F2")
            elif label.startswith("实际") and isinstance(val, (int, float)):
                ideal = ideal_remaining[j - 2]
                if val <= ideal:
                    c.fill = PatternFill("solid", fgColor="C6EFCE")
                elif val - ideal <= 15:
                    c.fill = PatternFill("solid", fgColor="FFEB9C")
                else:
                    c.fill = PatternFill("solid", fgColor="FFC7CE")

    note_row = row_hdr + 1 + len(summary_rows)
    ws.cell(note_row, 1, "当日备注").font = Font(bold=True, italic=True)
    ws.cell(note_row, 1).alignment = center
    for j, note in enumerate(day_notes, start=3):
        c = ws.cell(note_row, j, note)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=9, italic=True, color="666666")

    task_hdr = note_row + 2
    ws.cell(task_hdr, 1, "任务明细（理想时拆分）").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=task_hdr, start_column=1, end_row=task_hdr, end_column=last_col)

    task_col_hdr = task_hdr + 1
    task_headers = ["编号", "任务", "理想时(h)", "负责人", "完成日", "状态", "版本"]
    for col, text in enumerate(task_headers, start=1):
        cell = ws.cell(task_col_hdr, col, text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for i, (tid, name, hours, owner, done_day, status, version) in enumerate(tasks, start=task_col_hdr + 1):
        day_label = days[done_day][0] if done_day < len(days) else "-"
        row_vals = [tid, name, hours, owner, day_label, status, version]
        for col, val in enumerate(row_vals, start=1):
            c = ws.cell(i, col, val)
            c.alignment = center if col != 2 else Alignment(wrap_text=True, vertical="center")
            c.border = border
            if col == 6:
                c.fill = _status_fill(status)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    for col in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    _add_burndown_chart_sheet(wb)

    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ["OnlyFriends 燃尽图使用说明", ""],
        ["", ""],
        ["1. 统计范围", "本表覆盖第 0-3 天（2026-06-26 至 2026-06-29）v0.0 启动与规划阶段"],
        ["2. 理想时单位", "以「理想时（小时）」计量，1 理想时 ≈ 1 人专注工作 1 小时"],
        ["3. 理想线", "自期初总理想时按日程均匀燃尽至 0"],
        ["4. 实际线", "每日站会后根据任务完成情况更新剩余工时"],
        ["5. 更新频率", "每日站会结束后由 PO 或 Scrum Master 更新"],
        ["6. 预警规则", "实际剩余连续 2 天高于理想线 15h 以上时，在回顾会讨论范围或分工调整"],
        ["", ""],
        ["7. 日程与燃尽关联", ""],
        ["日程", "版本", "典型燃尽变化"],
        ["第0天 06-26", "启动", "完成团队规范，燃尽缓慢启动"],
        ["第1天 06-27", "v0.0", "文档类任务集中完成，燃尽加速"],
        ["第2天 06-28", "v0.0", "环境搭建遇阻，实际线高于理想线"],
        ["第3天 06-29", "v0.1 启动", "迭代计划与任务拆分完成，剩余 10h 转入首次迭代开发"],
        ["", ""],
        ["8. v0.1 后续", "自 2026-06-30 起应新建 v0.1 迭代燃尽图，跟踪主链路开发剩余工时"],
    ]
    for r, row in enumerate(instructions, 1):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(r, c, val)
            if r in (1, 7):
                cell.font = Font(bold=True, size=12 if r == 1 else 11)
            if r == 8:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E7E6E6")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 60

    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Size: {os.path.getsize(OUT_PATH)} bytes")


if __name__ == "__main__":
    main()
