# -*- coding: utf-8 -*-
"""Fill the Scrum import template from backend/docs/user-stories.md."""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
USER_STORIES_MD = ROOT / "backend" / "docs" / "user-stories.md"
TEMPLATE = ROOT / "scripts" / "scrum" / "importScrumModule.xlsx"
SHEET_NAME = "工作项列表"

PRIORITY_MAP = {
    "P0": ("高", "关键"),
    "P1": ("中", "重要"),
    "P2": ("低", "一般"),
}

VERSION_MAP = {
    "v0.0": "OnlyFriends_v0.0",
    "v0.1": "OnlyFriends_v0.1",
    "v0.5": "OnlyFriends_v0.5",
    "v1.0": "OnlyFriends_v1.0",
    "v1.1": "OnlyFriends_v1.1",
}

ITERATION_MAP = {
    "v0.0": "迭代0-启动与规划",
    "v0.1": "迭代1-首次迭代MVP",
    "v0.5": "迭代2-第二次迭代",
    "v1.0": "迭代3-最终验收",
    "v1.1": "迭代4-后续增强",
}

DATE_RANGE = {
    "v0.0": ("2026-06-27", "2026-06-28"),
    "v0.1": ("2026-06-29", "2026-07-01"),
    "v0.5": ("2026-07-03", "2026-07-06"),
    "v1.0": ("2026-07-07", "2026-07-07"),
    "v1.1": ("2026-07-08", "2026-07-21"),
}

MODULE_MAP = {
    "EPIC-01": "OnlyFriends/用户服务",
    "EPIC-02": "OnlyFriends/活动服务",
    "EPIC-03": "OnlyFriends/社群服务",
    "EPIC-04": "OnlyFriends/即时通讯",
    "EPIC-05": "OnlyFriends/AI服务",
    "EPIC-06": "OnlyFriends/后台管理",
    "EPIC-07": "OnlyFriends/通知模块",
    "EPIC-08": "OnlyFriends/基础设施",
}

EPIC_IDS = {
    "EPIC-01": 10001,
    "EPIC-02": 10002,
    "EPIC-03": 10003,
    "EPIC-04": 10004,
    "EPIC-05": 10005,
    "EPIC-06": 10006,
    "EPIC-07": 10007,
    "EPIC-08": 10008,
}


@dataclass(frozen=True)
class Epic:
    code: str
    title: str
    description: str
    version: str


@dataclass(frozen=True)
class Story:
    code: str
    epic_code: str
    priority: str
    version: str
    strategy: str
    role: str
    want: str
    so_that: str


def normalize_version(version: str) -> str:
    version = version.strip()
    if version.endswith("+"):
        return version[:-1]
    return version


def strip_md(value: str) -> str:
    return re.sub(r"\*\*", "", value).strip()


def split_table_row(line: str) -> list[str]:
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def parse_story_sentence(sentence: str) -> tuple[str, str, str]:
    plain = strip_md(sentence)
    match = re.search(r"作为\s+(.+?)，我想要\s+(.+?)，以便于\s+(.+?)[。.]?$", plain)
    if not match:
        raise ValueError(f"Cannot parse story sentence: {sentence}")
    return tuple(part.strip() for part in match.groups())


def parse_user_stories(markdown: str) -> tuple[list[Epic], list[Story]]:
    epics: list[Epic] = []
    stories: list[Story] = []
    current_epic: Epic | None = None

    lines = markdown.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]

        epic_match = re.match(r"^# (EPIC-\d{2}) (.+)$", line)
        if epic_match:
            code, title = epic_match.groups()
            attrs: dict[str, str] = {}
            i += 1
            while i < len(lines):
                if lines[i].startswith("|") and not lines[i].startswith("|------"):
                    cells = split_table_row(lines[i])
                    if len(cells) >= 2 and cells[0] not in ("属性", "Story ID"):
                        attrs[cells[0]] = cells[1]
                if lines[i].startswith("| Story ID "):
                    break
                if i > 0 and lines[i].startswith("# EPIC-"):
                    break
                i += 1

            description = attrs.get("业务价值", "")
            epic_version = "v0.0" if code == "EPIC-08" else "v0.1" if code in ("EPIC-01", "EPIC-05", "EPIC-06") else "v0.5"
            current_epic = Epic(
                code=code,
                title=title,
                description=description,
                version=epic_version,
            )
            epics.append(current_epic)
            continue

        if current_epic and line.startswith("| US-"):
            cells = split_table_row(line)
            if len(cells) >= 5:
                story_code, priority, version, strategy, sentence = cells[:5]
                role, want, so_that = parse_story_sentence(sentence)
                stories.append(
                    Story(
                        code=story_code,
                        epic_code=current_epic.code,
                        priority=priority,
                        version=normalize_version(version),
                        strategy=strategy,
                        role=role,
                        want=want,
                        so_that=so_that,
                    )
                )

        i += 1

    return epics, stories


def story_description(story: Story) -> str:
    return (
        f"作为 {story.role}\n"
        f"我想要 {story.want}\n"
        f"以便于 {story.so_that}\n\n"
        f"优先级：{story.priority}\n"
        f"目标版本：{story.version}\n"
        f"10天处理策略：{story.strategy}"
    )


def build_row(
    item_type: str,
    title: str,
    item_id: int,
    parent_id: int | None,
    description: str,
    priority: str,
    version: str,
    module: str,
    tags: str = "",
) -> list[object | None]:
    version = normalize_version(version)
    pri, importance = PRIORITY_MAP[priority]
    start_date, end_date = DATE_RANGE[version]
    return [
        item_type,
        title,
        item_id,
        parent_id,
        None,
        ITERATION_MAP[version],
        module,
        pri,
        importance,
        "新建",
        description,
        tags,
        "功能",
        start_date,
        end_date,
        VERSION_MAP[version],
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ]


def clear_data_rows(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def main() -> None:
    output = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else TEMPLATE
    markdown = USER_STORIES_MD.read_text(encoding="utf-8")
    epics, stories = parse_user_stories(markdown)

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb[SHEET_NAME]
    clear_data_rows(ws)

    rows: list[list[object | None]] = []
    for epic in epics:
        rows.append(
            build_row(
                item_type="epic",
                title=f"{epic.code} {epic.title}",
                item_id=EPIC_IDS[epic.code],
                parent_id=None,
                description=epic.description,
                priority="P0" if epic.code in ("EPIC-01", "EPIC-02", "EPIC-05", "EPIC-06", "EPIC-08") else "P1",
                version=epic.version,
                module=MODULE_MAP[epic.code],
                tags=epic.code,
            )
        )

    for offset, story in enumerate(stories, start=1):
        parent_id = EPIC_IDS[story.epic_code]
        title = f"{story.code} {story.want}"
        tags = f"{story.code} {story.priority} {story.strategy}"
        rows.append(
            build_row(
                item_type="story",
                title=title,
                item_id=20000 + offset,
                parent_id=parent_id,
                description=story_description(story),
                priority=story.priority,
                version=story.version,
                module=MODULE_MAP[story.epic_code],
                tags=tags,
            )
        )

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output)
    print(f"Filled {len(rows)} work items ({len(epics)} epics + {len(stories)} stories) into {output}")


if __name__ == "__main__":
    main()
